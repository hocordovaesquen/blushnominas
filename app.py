import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
import plotly.express as px
import numpy as np

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(
    page_title="BLUSH - Sistema de Comisiones",
    page_icon="💇‍♀️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- ESTILOS CSS ---
st.markdown("""
<style>
    .main-header {
        font-family: 'Arial', sans-serif;
        font-size: 2.5rem;
        font-weight: 700;
        color: #FFFFFF;
        text-align: center;
        padding: 1.5rem;
        background: linear-gradient(90deg, #E91E63 0%, #FF4081 100%);
        border-radius: 15px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 15px rgba(233, 30, 99, 0.3);
    }
    .stButton>button {
        background-color: #E91E63;
        color: white;
        font-weight: 600;
        border-radius: 8px;
        padding: 0.6rem 2rem;
        border: none;
        transition: all 0.3s ease;
    }
    .stButton>button:hover {
        background-color: #C2185B;
        transform: scale(1.02);
    }
    .metric-card {
        background-color: #FFFFFF;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        text-align: center;
        border-left: 5px solid #E91E63;
    }
</style>
""", unsafe_allow_html=True)

# --- FUNCIONES DE PROCESAMIENTO ---
@st.cache_data(ttl=3600)
def load_data(file):
    try:
        # 1. Leer primero sin encabezados para encontrar la fila real de títulos
        df_raw = pd.read_excel(file, header=None)
        
        header_idx = -1
        # Buscamos la fila que contenga "FECHA", "PRODUCTO" y "TOTAL"
        for i, row in df_raw.head(20).iterrows():
            row_str = " ".join(row.astype(str).values).upper()
            if "FECHA" in row_str and "PRODUCTO" in row_str and "TOTAL" in row_str:
                header_idx = i
                break
        
        if header_idx == -1:
            st.error("❌ No se encontró la fila de encabezados. Verifica que el Excel tenga columnas 'FECHA', 'PRODUCTO / SERVICIO' y 'TOTAL'.")
            return None

        # 2. Recargar con el encabezado correcto
        df = pd.read_excel(file, header=header_idx)
        
        # 3. Limpieza estricta de columnas
        df.columns = df.columns.str.strip().str.upper() # Todo a mayúsculas y sin espacios extra
        
        # Mapeo inteligente de columnas (por si cambian ligeramente)
        col_map = {}
        for col in df.columns:
            if "PRODUCTO" in col: col_map[col] = 'PRODUCTO / SERVICIO'
            elif "EMPLEADO" in col: col_map[col] = 'EMPLEADO'
            elif "TOTAL" in col and "COMP" not in col: col_map[col] = 'TOTAL' # Evitar TOTAL COMPROBANTE
            elif "TV" in col: col_map[col] = 'TV' # Columna de validación

        df = df.rename(columns=col_map)

        # 4. FILTROS DE SEGURIDAD (Aquí está la clave para evitar "más productos")
        
        # A) Eliminar filas totalmente vacías en Producto
        df = df.dropna(subset=['PRODUCTO / SERVICIO'])
        
        # B) Eliminar filas que son basura (líneas divisorias, espacios vacíos que parecen texto)
        df = df[df['PRODUCTO / SERVICIO'].astype(str).str.strip() != '']
        df = df[df['PRODUCTO / SERVICIO'].astype(str).str.lower() != 'nan']
        
        # C) Eliminar filas que sean repeticiones de encabezados o Totales del reporte
        trash_words = ['TOTAL', 'SUBTOTAL', 'RESUMEN', 'PRODUCTO / SERVICIO', 'REGISTRO VENTA DETALLE']
        df = df[~df['PRODUCTO / SERVICIO'].astype(str).str.upper().isin(trash_words)]

        # D) Filtrar solo Ventas Validas (Si existe columna TV, filtrar "V")
        if 'TV' in df.columns:
            df = df[df['TV'].astype(str).str.upper() == 'V']

        # 5. Conversión de tipos
        # Convertir TOTAL a numérico, forzando errores a NaN y luego a 0
        df['TOTAL'] = pd.to_numeric(df['TOTAL'], errors='coerce').fillna(0)
        
        # 6. Lógica de rellenado (Fillna) SOLO para fechas/empleados si es necesario,
        # pero en este reporte el empleado está en la misma línea del producto, así que cuidado.
        if 'EMPLEADO' in df.columns:
            df['EMPLEADO'] = df['EMPLEADO'].fillna('Sin Asignar')
            
        return df

    except Exception as e:
        st.error(f"Error al procesar el archivo: {str(e)}")
        return None

def procesar_nomina(df, comision_base):
    # Agrupar por Empleado
    resumen = df.groupby('EMPLEADO').agg(
        TOTAL_PRODUCCION=('TOTAL', 'sum'),
        CONTEO_SERVICIOS=('PRODUCTO / SERVICIO', 'count')
    ).reset_index()
    
    # Calcular comisión
    resumen['PORCENTAJE'] = comision_base / 100
    resumen['TOTAL_COMISION'] = resumen['TOTAL_PRODUCCION'] * resumen['PORCENTAJE']
    
    # Calcular participación
    total_global = resumen['TOTAL_PRODUCCION'].sum()
    if total_global > 0:
        resumen['PARTICIPACION'] = resumen['TOTAL_PRODUCCION'] / total_global
    else:
        resumen['PARTICIPACION'] = 0
        
    return resumen.sort_values('TOTAL_PRODUCCION', ascending=False)

def crear_excel_con_formulas(df_detalle, df_resumen):
    output = io.BytesIO()
    wb = Workbook()
    
    # --- ESTILOS EXCEL ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
    currency_format = '_("S/"* #,##0.00_);_("S/"* (#,##0.00);_("S/"* "-"??_);_(@_)'
    center_align = Alignment(horizontal='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # --- HOJA 1: RESUMEN ---
    ws1 = wb.active
    ws1.title = "Resumen Nómina"
    
    headers_resumen = ["EMPLEADO", "CANT. SERVICIOS", "VENTA TOTAL (S/)", "% COMISIÓN", "PAGO COMISIÓN (S/)", "% PARTICIPACIÓN"]
    ws1.append(headers_resumen)
    
    for col_num, header in enumerate(headers_resumen, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for i, row in df_resumen.iterrows():
        r = i + 2
        ws1.cell(row=r, column=1, value=row['EMPLEADO']).border = border
        ws1.cell(row=r, column=2, value=row['CONTEO_SERVICIOS']).border = border
        
        # Venta Total
        c3 = ws1.cell(row=r, column=3, value=row['TOTAL_PRODUCCION'])
        c3.number_format = currency_format
        c3.border = border
        
        # % Comisión (Input manual posible si se descarga)
        c4 = ws1.cell(row=r, column=4, value=row['PORCENTAJE'])
        c4.number_format = '0%'
        c4.border = border
        
        # Fórmula Comisión: Venta * %
        c5 = ws1.cell(row=r, column=5, value=f"=C{r}*D{r}")
        c5.number_format = currency_format
        c5.font = Font(bold=True)
        c5.border = border
        
        # Participación
        c6 = ws1.cell(row=r, column=6, value=row['PARTICIPACION'])
        c6.number_format = '0.0%'
        c6.border = border

    # Ajustar ancho columnas
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws1.column_dimensions[column].width = max_length + 2

    # --- HOJA 2: DETALLE ---
    ws2 = wb.create_sheet("Detalle Operaciones")
    
    # Seleccionar columnas clave para el detalle
    cols_detalle = ['FECHA', 'EMPLEADO', 'PRODUCTO / SERVICIO', 'TOTAL', 'CLIENTE']
    # Asegurarnos que existen en el DF
    cols_exportar = [c for c in cols_detalle if c in df_detalle.columns]
    
    # Escribir encabezados
    for col_num, header in enumerate(cols_exportar, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    # Escribir datos
    for r_idx, row in enumerate(df_detalle[cols_exportar].itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if c_idx == 4: # Asumiendo que TOTAL es la 4ta columna
                cell.number_format = currency_format

    wb.save(output)
    return output.getvalue()

# --- INTERFAZ DE USUARIO ---
st.markdown('<div class="main-header">💇‍♀️ BLUSH - Cálculo de Nómina</div>', unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.image("https://img.freepik.com/free-vector/hair-salon-logo-design_23-2149684162.jpg", width=150) # Placeholder o logo
    st.header("⚙️ Configuración")
    comision_input = st.slider("Porcentaje de Comisión General", 0, 100, 50, format="%d%%")
    st.info("ℹ️ Sube el archivo Excel 'Listado de Registro Ventas' para calcular.")

uploaded_file = st.file_uploader("📂 Cargar Excel de Ventas", type=['xlsx', 'xls'])

if uploaded_file:
    df = load_data(uploaded_file)
    
    if df is not None:
        # Procesar datos
        resumen = procesar_nomina(df, comision_input)
        
        # Métricas Top
        st.markdown("### 📊 Resumen del Periodo")
        col1, col2, col3 = st.columns(3)
        total_v = resumen['TOTAL_PRODUCCION'].sum()
        total_c = resumen['TOTAL_COMISION'].sum()
        
        # Estilo de métricas
        col1.metric("Venta Total Salón", f"S/ {total_v:,.2f}")
        col2.metric("Comisión a Pagar", f"S/ {total_c:,.2f}", delta=f"{comision_input}% Base")
        col3.metric("Total Servicios Realizados", int(resumen['CONTEO_SERVICIOS'].sum()))
        
        st.markdown("---")
        
        # Tabs para ver resumen y detalle
        tab1, tab2 = st.tabs(["💰 Nómina por Empleado", "📝 Detalle de Servicios"])
        
        with tab1:
            st.markdown("#### 🏆 Tabla de Comisiones")
            # Formatear para mostrar en pantalla
            st.dataframe(
                resumen.style.format({
                    'TOTAL_PRODUCCION': 'S/ {:,.2f}',
                    'TOTAL_COMISION': 'S/ {:,.2f}', 
                    'PARTICIPACION': '{:.1%}'
                }).background_gradient(cmap='Reds', subset=['TOTAL_PRODUCCION']),
                use_container_width=True,
                height=400
            )
            
            # Gráfico
            fig = px.bar(resumen, x='EMPLEADO', y='TOTAL_PRODUCCION', 
                         text_auto='.2s', title="Producción por Estilista",
                         color='TOTAL_PRODUCCION', color_continuous_scale='pinkyl')
            fig.update_layout(xaxis_title="", yaxis_title="Ventas (S/)")
            st.plotly_chart(fig, use_container_width=True)

        with tab2:
            st.markdown("#### 🔍 Auditoría de Productos Detectados")
            st.warning(f"Se han detectado {len(df)} servicios válidos después de filtrar anulados y filas vacías.")
            st.dataframe(df, use_container_width=True)

        # Botón de Descarga
        st.markdown("### 👇 Descarga Final")
        excel_data = crear_excel_con_formulas(df, resumen)
        
        st.download_button(
            label="📥 DESCARGAR NOMINA FINAL (EXCEL)",
            data=excel_data,
            file_name=f"Nomina_Blush_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    
else:
    st.info("👆 Esperando archivo... Por favor carga el Excel arriba.")